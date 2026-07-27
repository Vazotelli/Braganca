# -*- coding: utf-8 -*-
"""
gerar_seed.py
=============
Le o Excel `Mediçoes e custos.xlsx` (no OneDrive, so-leitura) e escreve o seed
inicial da app em /obra/js/dados.js.

Corre outra vez sempre que atualizares o Excel:
    python scripts/gerar_seed.py

Folhas lidas: Medições, Equipamentos, Orçamento construção, Materiais,
Gantt, Pagamentos MO.

Correcoes aplicadas na migracao (problemas conhecidos do Excel — nao propagar):
  #1 "Massas de cimento" com #VALUE! (qtd '?')  -> materiais com qtd=null,
     estado 'por_orcamentar'.
  #2 Pavimentos Apt.D com "Que fazer aqui?"      -> medicao estado 'por_decidir'.
  #3 Capitulos "Gesso e Pladur" e "Tetos falsos" vazios -> linha placeholder
     'por_orcamentar' de ~6.500 EUR (estimativa conhecida).
  #4 Equipamentos com 2 precos (novo/usado)      -> precoNovo e precoUsado.
  #5 Materiais com credito/devolucao (-150/-225) -> preserva negativos.
"""

from pathlib import Path
from datetime import date, datetime
import json
import sys

import openpyxl

# --------------------------------------------------------------------------
# CAMINHOS  (muda aqui se o OneDrive mudar de local)
# --------------------------------------------------------------------------
XLSX = Path(
    r"C:\Users\João Ribeiro\OneDrive - Kaizen Institute Ltd. (KIAG)"
    r"\Documents\11.Bragança\Mediçoes e custos.xlsx"
)
RAIZ_REPO = Path(__file__).resolve().parent.parent
DESTINO_JS = RAIZ_REPO / "obra" / "js" / "dados.js"

IVA_MATERIAL_PADRAO = 0.23   # materiais a retalho
IVA_MO_PADRAO = 0.06         # empreitada / reabilitacao


# --------------------------------------------------------------------------
# AUXILIARES
# --------------------------------------------------------------------------
def num(v):
    """Devolve float se for numero (nao texto/#VALUE!/None), senao None."""
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    return None


def texto(v):
    return "" if v is None else str(v).strip()


def iso(v):
    """datetime -> 'YYYY-MM-DD'; senao None."""
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    return None


# --------------------------------------------------------------------------
# MATERIAIS
# --------------------------------------------------------------------------
def ler_materiais(wb):
    ws = wb["Materiais"]
    materiais = []
    capitulo = ""
    for fila in ws.iter_rows(min_row=3, values_only=True):
        adj, esp, artigo, qtd, un, pu, total, iva, obs = (list(fila) + [None] * 9)[:9]
        c_artigo = texto(artigo)
        c_col_a = texto(adj)
        # cabecalho de capitulo: "── NOME ──" (vem na coluna A da folha)
        marcador = (c_col_a if c_col_a.startswith("──")
                    else c_artigo if c_artigo.startswith("──") else None)
        if marcador:
            capitulo = marcador.strip("─ ").strip()
            continue
        # linhas de total no fim da folha
        if c_artigo.upper().startswith("TOTAL"):
            continue
        # linha de material valida = tem artigo (a especialidade as vezes falta)
        if not c_artigo:
            continue
        g_total = num(total)
        g_iva = num(iva)
        g_iva = g_iva if (g_iva is not None and g_iva > 0) else None
        taxa = (round(g_iva / g_total, 2)
                if (g_iva and g_total and g_total > 0) else IVA_MATERIAL_PADRAO)
        materiais.append({
            "id": f"mat_{len(materiais)+1:03d}",
            "capitulo": capitulo,
            "especialidade": texto(esp) or capitulo,
            "artigo": c_artigo,
            "qtd": num(qtd),
            "unidade": texto(un),
            "precoUnit": round(num(pu), 5) if num(pu) is not None else None,
            "total": g_total,
            "iva": g_iva,
            "ivaAtivo": g_iva is not None,   # IVA aplicado nesta linha (como no Excel)
            "taxaIva": taxa,
            "adjudicado": adj is True,
            "fornecedor": None,
            "observacoes": texto(obs),
            "estado": "ok",
        })
    return materiais


def placeholders_gesso_tetos(materiais):
    """#3 — capitulos vazios: cria placeholders por_orcamentar (~6.500 EUR)."""
    tem = {m["capitulo"] for m in materiais}
    extras = []
    if "Gesso e Pladur" in tem and not any(
            m["capitulo"] == "Gesso e Pladur" and m["artigo"] for m in materiais
            if m["artigo"] not in ("", None)):
        pass  # os capitulos existem como cabecalho mas sem linhas com artigo
    # As folhas tem os cabecalhos mas linhas em branco (sem artigo) -> ja
    # foram ignoradas em ler_materiais. Criamos 1 placeholder por capitulo.
    for cap, estimativa, nota in [
        ("Gesso e Pladur", 3250, "~320 m² paredes pladur (estimativa)"),
        ("Tetos falsos", 3250, "~120 m² tetos falsos (estimativa)"),
    ]:
        extras.append({
            "id": f"mat_ph_{cap.split()[0].lower()}",
            "capitulo": cap,
            "especialidade": cap,
            "artigo": "Estimativa global (a orçamentar)",
            "qtd": None,
            "unidade": "vg",
            "precoUnit": None,
            "total": None,               # nao conta para estimado/gasto
            "iva": None,
            "ivaAtivo": False,
            "taxaIva": IVA_MATERIAL_PADRAO,
            "adjudicado": False,
            "fornecedor": None,
            "observacoes": f"{nota}. Estimativa ~{estimativa} EUR + IVA.",
            "estado": "por_orcamentar",
            "estimativaInformativa": estimativa,
        })
    return extras


def massas_cimento(wb):
    """#1 — 4 linhas 'Massas de cimento' com #VALUE! na folha Medições."""
    ws = wb["Medições"]
    linhas = []
    for fila in ws.iter_rows(min_row=1, values_only=True):
        vals = list(fila) + [None] * 18
        if texto(vals[2]) == "Massas de cimento":
            pu = num(vals[10])  # coluna K = P.U.
            linhas.append({
                "id": f"mat_massa_{len(linhas)+1}",
                "capitulo": "ALVENARIA",
                "especialidade": "Alvenaria",
                "artigo": "Massas de cimento",
                "qtd": None,                 # era '?' no Excel
                "unidade": "sc",
                "precoUnit": pu,
                "total": None,               # #VALUE! -> nao conta
                "iva": None,
                "ivaAtivo": False,
                "taxaIva": IVA_MATERIAL_PADRAO,
                "adjudicado": False,
                "fornecedor": "Leroy Merlin",
                "observacoes": "Quantidade por orçamentar.",
                "estado": "por_orcamentar",
            })
    return linhas


# --------------------------------------------------------------------------
# MEDICOES  (totais por capitulo, so leitura)
# --------------------------------------------------------------------------
def _acha(ws, col_a, col_c):
    """Devolve a fila (tuple) onde coluna A==col_a e coluna C==col_c."""
    for fila in ws.iter_rows(min_row=1, values_only=True):
        vals = list(fila) + [None] * 18
        if texto(vals[0]) == col_a and texto(vals[2]) == col_c:
            return vals
    return None


def ler_medicoes(wb):
    ws = wb["Medições"]
    medicoes = []

    fp = _acha(ws, "Paredes", "Paredes")
    if fp:
        medicoes.append({
            "capitulo": "Paredes", "quantidade": num(fp[8]), "unidade": "m2",
            "precoUnit": num(fp[10]),
            "notas": "Parede central 78,09 m² + Apt.E 18,52 + Apt.D 20,91",
            "estado": "ok",
        })

    fa = _acha(ws, "Azulejo", "Azulejo")
    if fa:
        medicoes.append({
            "capitulo": "Azulejo", "quantidade": num(fa[8]), "unidade": "m2",
            "precoUnit": num(fa[10]),
            "notas": "Inclui 6,4% de quebra. IS Esq 27,54 + IS Dir 29,30",
            "estado": "ok",
        })

    pe = _acha(ws, "Pavimentos", "Apartamento E")
    pd = _acha(ws, "Pavimentos", "Apartamento D")
    if pe and pd:
        q = (num(pe[8]) or 0) + (num(pd[8]) or 0)
        medicoes.append({
            "capitulo": "Pavimentos", "quantidade": round(q, 4), "unidade": "m2",
            "precoUnit": None,
            "notas": "Apt.E 72,41 (int. 55,05) + Apt.D 70,99 (int. 53,74). "
                     "Apt.D por decidir (Que fazer aqui?).",
            "estado": "por_decidir",   # #2
        })

    lr = _acha(ws, "Paredes", "Lã de Rocha")
    if lr:
        medicoes.append({
            "capitulo": "Lã de rocha", "quantidade": num(lr[8]), "unidade": "m2",
            "precoUnit": None, "notas": "", "estado": "ok",
        })

    return medicoes


# --------------------------------------------------------------------------
# ESPECIALIDADES  (folha Pagamentos MO)
# --------------------------------------------------------------------------
COLUNAS_MO = [
    # (col_idx 0-based, id, nome)
    (1, "arquiteto",         "Arquiteto — Harry Martins"),
    (2, "trolhas",           "Trolhas — RC Remodelações"),
    (3, "peitoris",          "Peitoris"),
    (4, "caixilharia",       "Caixilharia"),
    (5, "pichelaria",        "Pichelaria"),
    (6, "carpinteiro",       "Carpinteiro"),
    (7, "trolhas_varandas",  "Trolhas — Varandas + café"),
]


def _metodo_desc(txt):
    linhas = [l.strip() for l in texto(txt).split("\n") if l.strip()]
    metodo = "Transferência"
    desc = []
    for l in linhas:
        low = l.lower()
        if low.startswith("transfer") or low.startswith("dinheiro"):
            metodo = "Transferência" if low.startswith("transfer") else "Dinheiro"
        else:
            desc.append(l)
    return metodo, " + ".join(desc)


def ler_especialidades(wb):
    ws = wb["Pagamentos MO"]
    filas = list(ws.iter_rows(min_row=1, values_only=True))

    # localizar linha Total e % paga
    total_row = pct_row = None
    for f in filas:
        a = texto(f[0]).lower()
        if a == "total":
            total_row = list(f) + [None] * 8
        elif a in ("% paga", "% pago"):
            pct_row = list(f) + [None] * 8

    # recolher pares (data, metodo) -> marcos por coluna
    marcos_por_col = {c: [] for c, _, _ in COLUNAS_MO}
    for i, f in enumerate(filas):
        d = iso(f[0])
        if not d:
            continue
        metodo_row = list(filas[i + 1]) + [None] * 8 if i + 1 < len(filas) else [None] * 8
        for c, _id, _nome in COLUNAS_MO:
            valor = num((list(f) + [None] * 8)[c])
            if valor is None:
                continue
            metodo, desc = _metodo_desc(metodo_row[c])
            marcos_por_col[c].append({"valor": valor, "data": d,
                                       "metodo": metodo, "descricao": desc})

    especialidades = []
    for c, _id, _nome in COLUNAS_MO:
        total = num(total_row[c]) if total_row else None
        pct = num(pct_row[c]) if pct_row else 0.0
        alvo_pago = (pct or 0) * (total or 0)
        marcos = sorted(marcos_por_col[c], key=lambda m: m["data"])
        acumulado = 0.0
        saida = []
        for j, m in enumerate(marcos):
            pago = acumulado + m["valor"] <= alvo_pago + 0.5
            if pago:
                acumulado += m["valor"]
            saida.append({
                "id": f"m{j+1}",
                "descricao": m["descricao"],
                "valor": m["valor"],
                "dataPrevista": m["data"],
                "dataPaga": m["data"] if pago else None,
                "metodo": m["metodo"],
            })
        especialidades.append({
            "id": _id, "nome": _nome,
            "totalContratado": total, "taxaIva": IVA_MO_PADRAO,
            "marcos": saida,
        })
    return especialidades


# --------------------------------------------------------------------------
# FASES  (folha Gantt)
# --------------------------------------------------------------------------
def ler_fases(wb):
    ws = wb["Gantt"]
    fases = []
    fase_atual = None
    for f in ws.iter_rows(min_row=3, values_only=True):
        vals = list(f) + [None] * 30
        label = texto(vals[0])
        if not label:
            continue
        # marcas: colunas B..AD (indices 1..29 0-based) -> semana = idx + 23
        marcas = vals[1:30]

        partes = label.split("|", 1)
        e_fase = len(partes) == 2 and partes[0].strip().isdigit()
        if e_fase:
            numero = partes[0].strip().zfill(2)
            nome = partes[1].strip()
            semanas = [i + 23 for i, v in enumerate(marcas)
                       if texto(v).upper() == "A"]
            fase_atual = {
                "id": numero, "numero": numero, "nome": nome,
                "semanaInicio": min(semanas) if semanas else None,
                "semanaFim": max(semanas) if semanas else None,
                "estado": "nao_iniciada",
                "dependeDe": [],            # definido no Passo 5
                "plantasAssociadas": [],    # definido no Passo 5
                "subtarefas": [],
            }
            fases.append(fase_atual)
        elif fase_atual is not None:
            # subtarefa: nome em A, marcas 'X' -> semana = idx + 23
            semanas = [i + 23 for i, v in enumerate(marcas)
                       if texto(v).upper() == "X"]
            if not semanas:
                continue
            fase_atual["subtarefas"].append({
                "nome": label,
                "semanaInicio": min(semanas),
                "semanaFim": max(semanas),
                "concluida": False,
            })
    _aplicar_dependencias(fases)
    return fases


# Grafo de dependencias entre fases (logica de obra) e plantas associadas.
# Nao vem do Excel — definido aqui para o ecra Cronograma.
_DEPENDENCIAS = {
    "01": [], "02": ["01"], "03": ["01", "02"], "04": ["03"],
    "05": ["03", "04"], "06": ["03", "04", "05"], "07": ["05", "06"],
    "08": ["03"], "09": ["06"], "10": ["06", "09"], "11": ["06"],
    "12": ["06", "07"], "13": ["04", "10"], "14": ["06"],
    "15": ["06", "07", "10"], "16": ["12", "13", "14", "15"],
}
_PLANTAS_POR_FASE = {
    "01": ["01_demolicoes"],
    "03": ["02_alvenarias"],
    "04": ["03_saneamento_agua"],
    "05": ["05_iluminacao", "06_interruptores", "08_tomadas"],
    "08": ["07_peitoris"],
    "12": ["05_iluminacao", "06_interruptores", "08_tomadas"],
    "13": ["03_saneamento_agua"],
    "14": ["04_carpintarias"],
}


def _aplicar_dependencias(fases):
    for f in fases:
        f["dependeDe"] = _DEPENDENCIAS.get(f["numero"], [])
        f["plantasAssociadas"] = _PLANTAS_POR_FASE.get(f["numero"], [])


# --------------------------------------------------------------------------
# EQUIPAMENTOS  (#4 dois precos) e ORCAMENTO CONSTRUCAO
# --------------------------------------------------------------------------
def ler_equipamentos(wb):
    ws = wb["Equipamentos"]
    equip = []
    for fila in ws.iter_rows(min_row=2, values_only=True):
        nome, qtd, un, novo, usado, coment = (list(fila) + [None] * 6)[:6]
        if not texto(nome) or num(novo) is None:
            continue
        equip.append({
            "nome": texto(nome), "qtd": num(qtd), "unidade": texto(un),
            "precoNovo": num(novo), "precoUsado": num(usado),
            "comentario": texto(coment),
        })
    return equip


def ler_orcamento_construcao(wb):
    ws = wb["Orçamento construção"]
    linhas = []
    for fila in ws.iter_rows(min_row=3, values_only=True):
        esp, mo, mat, quem, met, ini, fim, fase, obs = (list(fila) + [None] * 9)[:9]
        if not texto(esp) or texto(esp).upper() == "TOTAL":
            continue
        linhas.append({
            "especialidade": texto(esp), "moEstimada": num(mo),
            "materiais": texto(mat), "quemFornece": texto(quem),
            "observacoes": texto(obs),
        })
    return linhas


# --------------------------------------------------------------------------
# TAREFAS ADMINISTRATIVAS  (nao existem no Excel — seed do brief 6.4)
# --------------------------------------------------------------------------
def tarefas_administrativas():
    return [
        {"id": "adm_01", "nome": "Licenciamento — mudança de uso (Câmara de Bragança)",
         "responsavel": "Harry Martins", "dataSubmissao": None, "dataAprovacao": None,
         "estado": "nao_iniciada", "bloqueia": ["adm_03", "16"]},
        {"id": "adm_02", "nome": "Constituição de propriedade horizontal",
         "responsavel": "Harry Martins", "dataSubmissao": None, "dataAprovacao": None,
         "estado": "nao_iniciada", "bloqueia": ["venda"]},
        {"id": "adm_03", "nome": "Contadores definitivos E-REDES",
         "responsavel": "E-REDES", "dataSubmissao": None, "dataAprovacao": None,
         "estado": "nao_iniciada", "bloqueia": []},
        {"id": "adm_04", "nome": "Certificação elétrica DGEG",
         "responsavel": "Eletricista", "dataSubmissao": None, "dataAprovacao": None,
         "estado": "nao_iniciada", "bloqueia": []},
        {"id": "adm_05", "nome": "Certificado energético (CDE)",
         "responsavel": "Perito qualificado", "dataSubmissao": None, "dataAprovacao": None,
         "estado": "nao_iniciada", "bloqueia": ["venda"]},
    ]


# --------------------------------------------------------------------------
# PRINCIPAL
# --------------------------------------------------------------------------
def acoes_iniciais(fases, admin):
    """Plano de acoes: cada acao = descricao + data de fim combinada em obra.
    Sementeia a partir das 16 fases (data de fim = sexta-feira da semana ISO) e
    das tarefas administrativas (sem data — o Joao define). Tudo editavel/apagavel."""
    acoes = []
    for f in fases:
        dfim = None
        if f.get("semanaFim"):
            try:
                dfim = date.fromisocalendar(2026, int(f["semanaFim"]), 5).isoformat()
            except ValueError:
                dfim = None
        acoes.append({
            "id": f"acao_f{f['numero']}",
            "descricao": f"{f['numero']} — {f['nome']}",
            "dataFim": dfim, "feito": False,
        })
    for t in admin:
        acoes.append({
            "id": f"acao_{t['id']}", "descricao": t["nome"],
            "dataFim": None, "feito": False,
        })
    return acoes


def negocio_inicial():
    """Investimento e venda (Joao/Joana). Nao vem do Excel — valores-base
    editaveis na app. Regra: Joana investe 1/3 do total, com teto de 20.000.
    Lucro do Joao = %investimento dele + metade da %investimento da Joana
    (porque e' ele que faz a obra)."""
    return {
        "investimentoTotal": 80000,   # editavel
        "precoVendaE": 130000,        # editavel (Apartamento Esquerdo)
        "precoVendaD": 120000,        # editavel (Apartamento Direito)
        "taxaImpostos": 0.20,         # editavel
    }


def main():
    if not XLSX.exists():
        print(f"ERRO: nao encontro o Excel em:\n  {XLSX}")
        return 1
    wb = openpyxl.load_workbook(XLSX, data_only=True)

    materiais = ler_materiais(wb)
    materiais += massas_cimento(wb)          # #1
    materiais += placeholders_gesso_tetos(materiais)  # #3
    medicoes = ler_medicoes(wb)              # #2 tratado dentro
    especialidades = ler_especialidades(wb)
    fases = ler_fases(wb)
    equipamentos = ler_equipamentos(wb)      # #4
    orcamento = ler_orcamento_construcao(wb)

    # --- totais de controlo (mesmas formulas do Excel) ---
    # Estimado  = SUMIF(TOTAL>0) + SUMIF(IVA>0)     (exclui creditos negativos)
    # Gasto     = SUMIF(adj, TOTAL) + SUMIF(adj, IVA)   (com IVA)
    def pos(xs):
        return sum(x for x in xs if isinstance(x, (int, float)) and x > 0)

    est_semiva = pos(m["total"] for m in materiais)
    est_iva = pos(m.get("iva") for m in materiais)
    est_comiva = est_semiva + est_iva

    gasto_semiva = sum(m["total"] for m in materiais
                       if m["adjudicado"] and num(m["total"]))
    gasto_iva = sum(m["iva"] for m in materiais
                    if m["adjudicado"] and num(m.get("iva")))
    gasto_comiva = gasto_semiva + gasto_iva

    total_mo = sum(e["totalContratado"] for e in especialidades
                   if e["totalContratado"])
    pago_mo = sum(mk["valor"] for e in especialidades for mk in e["marcos"]
                  if mk["dataPaga"])
    total_geral = total_mo + est_comiva
    gasto_total = gasto_comiva + pago_mo

    totais = {
        "materiaisEstimadoSemIva": round(est_semiva, 2),
        "materiaisEstimado": round(est_comiva, 2),      # c/ IVA (controlo 24.763)
        "materiaisGastoSemIva": round(gasto_semiva, 2),
        "materiaisGasto": round(gasto_comiva, 2),       # c/ IVA (controlo 3.994)
        "moTotal": round(total_mo, 2),
        "moPago": round(pago_mo, 2),
        "totalGeral": round(total_geral, 2),
        "gastoTotal": round(gasto_total, 2),
        "percGasto": round(gasto_total / total_geral, 4) if total_geral else 0,
    }

    seed = {
        "meta": {
            "geradoEm": date.today().isoformat(),
            "fonte": XLSX.name,
            "moeda": "EUR",
            "totaisControlo": totais,
        },
        "materiais": materiais,
        "medicoes": medicoes,
        "especialidades": especialidades,
        "fases": fases,
        "tarefasAdministrativas": tarefas_administrativas(),
        "acoes": acoes_iniciais(fases, tarefas_administrativas()),
        "equipamentos": equipamentos,
        "orcamentoConstrucao": orcamento,
        "negocio": negocio_inicial(),
    }

    corpo = json.dumps(seed, ensure_ascii=False, indent=2)
    DESTINO_JS.parent.mkdir(parents=True, exist_ok=True)
    DESTINO_JS.write_text(
        "// GERADO POR scripts/gerar_seed.py — nao editar a mao.\n"
        "// Corre `python scripts/gerar_seed.py` para regenerar a partir do Excel.\n"
        f"export const SEED = {corpo};\n",
        encoding="utf-8",
    )

    # --- relatorio para validacao ---
    print(f"Escrito: {DESTINO_JS}")
    print(f"  materiais: {len(materiais)}  medicoes: {len(medicoes)}  "
          f"especialidades: {len(especialidades)}  fases: {len(fases)}")
    print(f"  subtarefas: {sum(len(f['subtarefas']) for f in fases)}  "
          f"equipamentos: {len(equipamentos)}  orcamento: {len(orcamento)}")
    print("\n=== TOTAIS DE CONTROLO (validar contra a seccao 4) ===")
    print(f"  Total MO (7 especialidades)     : {totais['moTotal']:>12,.2f}  (esperado 46.150)")
    print(f"  Materiais estimado  s/IVA       : {totais['materiaisEstimadoSemIva']:>12,.2f}")
    print(f"  Materiais estimado  c/IVA       : {totais['materiaisEstimado']:>12,.2f}  (esperado 24.763)")
    print(f"  TOTAL GERAL                     : {totais['totalGeral']:>12,.2f}  (esperado 70.913)")
    print(f"  Gasto materiais (adj) s/IVA     : {totais['materiaisGastoSemIva']:>12,.2f}")
    print(f"  Gasto materiais (adj) c/IVA     : {totais['materiaisGasto']:>12,.2f}  (esperado 3.994 = 16,1%)")
    print(f"  Pago mao de obra                : {totais['moPago']:>12,.2f}  (esperado 6.417)")
    print(f"  GASTO TOTAL ate a data          : {totais['gastoTotal']:>12,.2f}  (esperado 10.411 = 14,7%)")
    print(f"  % gasto                         : {totais['percGasto']*100:>11.2f}%  (esperado 14,7%)")
    print("\n  % pago por especialidade:")
    for e in especialidades:
        pago = sum(mk["valor"] for mk in e["marcos"] if mk["dataPaga"])
        tot = e["totalContratado"] or 0
        pct = (pago / tot * 100) if tot else 0
        prox = next((mk for mk in e["marcos"] if not mk["dataPaga"]), None)
        prox_txt = f"  proximo: {prox['valor']:.0f} EUR em {prox['dataPrevista']}" if prox else ""
        print(f"    {e['nome']:<32} {tot:>9,.0f}  {pct:>5.1f}%{prox_txt}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
